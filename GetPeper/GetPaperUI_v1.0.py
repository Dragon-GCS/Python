#!/usr/bin/env python3
# coding = utf-8
# Dragon's code
# vision = 1.0
# 20200825 完成UI主体框架
# 20200826 完成美化及主要按钮与功能测试
# 20200901 完成PubMed主要搜索功能与信息保存
# 20200908 完成ACS搜索功能


import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.font as tf
import urllib.request as ur
import GetFromPubMed as PM
import GetFromACS as ACS
import openpyxl as px
from bs4 import BeautifulSoup
from math import ceil
from os import mkdir, chdir
from os.path import exists


class Application:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('GetPaper_1.0')
        # 获取系统分辨率
        self.screenWidth = self.root.winfo_screenwidth()
        self.screenHeight = self.root.winfo_screenheight()
        # 限制窗口大小
        self.root.resizable(0, 0)
        # 设置字体
        self.ft = tf.Font(family='微软雅黑', size=13)
        # 边距
        tk.Label(self.root, width=1, height=1).grid(row=0, column=0)
        tk.Label(self.root, width=1, height=0).grid(row=5, column=12)
        # 数据存储
        self.title_list = []
        self.author_list = []
        self.publication_list = []
        self.date_list = []
        self.abstract_list = []
        self.doi_list = []
        self.web_list = []
        # 表头
        self.columns = ('Title', 'Author', 'Data', 'Publication', 'Abstract', 'DOI', 'Web Address')

        # 第一行
        # 所查询的数据库
        tk.Label(self.root, text='查询数据库：', font=self.ft).grid(row=1, column=1, sticky='e')
        self.datebase = ttk.Combobox(self.root, value=['PubMed', 'ACS', 'DataBase3 is coming'], font=self.ft)
        self.datebase.grid(row=1, column=2, columnspan=5, sticky='w')
        self.datebase.current(0)

        # 第二行
        # 搜索提示
        tk.Label(self.root, text='请输入需要查找的关键词：', font=self.ft).grid(row=2, column=1, sticky='e')
        # 关键词
        self.keywd = tk.Entry(self.root, font=self.ft)
        self.keywd.grid(row=2, column=2, columnspan=7, sticky='we')
        self.keywd.focus()  # 程序运行时光标默认出现在窗口中
        # 搜索键
        search_button = tk.Button(self.root, text='Search', command=self.search, width=8, font=self.ft)
        search_button.grid(row=2, column=9, columnspan=2, sticky='w', padx=5)
        search_button.bind('<Button-1>', self.search_tip)

        # 第三行
        tk.Label(self.root, text='共查找到：', font=self.ft).grid(row=3, column=1, sticky='e')
        tk.Label(self.root, text='', width=1, font=self.ft).grid(row=3, column=2, sticky='ew')
        tk.Label(self.root, text='篇文献', font=self.ft).grid(row=3, column=3, sticky='w')
        tk.Label(self.root, text='请输入需要获取的文献数量：', font=self.ft).grid(row=3, column=7, sticky='e')
        # 输入文献数量
        self.num = tk.Entry(self.root, width=6, font=self.ft)
        self.num.grid(row=3, column=8)
        # 开始获取文献信息
        tk.Button(self.root, text='Get Them', command=self.get_them, width=8, font=self.ft) \
            .grid(row=3, column=9, columnspan=2, sticky='w', padx=5)

        # 创建表格
        self.excel = ttk.Treeview(self.root, show='headings', columns=self.columns,
                                  height=int((self.screenHeight - 10) * 52 / 1920))
        self.excel.grid(row=4, column=1, columnspan=9, sticky='we')
        # 设置表格宽度与文字居中
        self.excel.column('Title', width=int(self.screenWidth / 8), anchor='center')
        self.excel.column('Author', width=int(self.screenWidth / 10), anchor='center')
        self.excel.column('Data', width=int(self.screenWidth / 20), anchor='center')
        self.excel.column('Publication', width=int(self.screenWidth / 16), anchor='center')
        self.excel.column('Abstract', width=int(self.screenWidth / 6), anchor='center')
        self.excel.column('DOI', width=int(self.screenWidth / 16), anchor='center')
        self.excel.column('Web Address', width=int(self.screenWidth / 16), anchor='center')
        # 设置表头
        for i in self.columns:
            self.excel.heading(i, text=i)
        # 单击打开文献详情
        self.excel.bind('<ButtonRelease-1>', self.open_abs)
        # 添加垂直滚动条
        vbar = ttk.Scrollbar(self.root, orient='vertical', command=self.excel.yview())
        self.excel.configure(yscrollcommand=vbar.set)
        vbar.grid(row=4, column=11, sticky='WNS')
        # 添加水平滚动条
        hbar = ttk.Scrollbar(self.root, orient='horizontal', command=self.excel.xview())
        self.excel.configure(xscrollcommand=hbar.set)
        hbar.grid(row=5, column=1, columnspan=10, sticky='NWE')

        # 保存、下载按钮
        menubar = tk.Menu(self.root)
        menubar.add_command(label='保存', command=self.file_save)
        menubar.add_command(label='下载', command=self.download)
        self.root['menu'] = menubar

    def run_it(self):
        self.root.mainloop()

    def search(self):
        if self.datebase.get() == 'PubMed':
            self.db = PM.GetFromPubmed()
        elif self.datebase.get() == 'ACS':
            self.db = ACS.GetFromACS()
        self.db.keyword = self.keywd.get()
        tk.Label(self.root, text=self.db.get_paper_max_num(), width=1, font=self.ft).grid(row=3, column=2, sticky='ew')

    def get_them(self):
        if self.datebase.get() == 'PubMed':
            new = tk.Toplevel()
            win = tk.Text(new)
            win.pack()

            self.db.paper_number = int(self.num.get())
            self.db.get_pmid()
            page = ceil(self.db.paper_number / 100)
            if page - 1:
                for i in range(2, page + 1):
                    self.db.get_paper_max_num(page=i)
                    self.db.get_pmid()
            for pmid in self.db.pmid_addr:
                self.db.count += 1
                win.insert('end', '正在获取第%i篇,共计%s篇\n' % (self.db.count, len(self.db.pmid_addr)))
                win.update()
                self.db.get_content(pmid)
            win.insert('end', '获取完成')
            win.update()
        elif self.datebase.get() == 'ACS':
            self.db.main(self.num.get())
        self.title_list, self.author_list, self.publication_list, self.date_list, self.abstract_list, self.doi_list, self.web_list = \
            [self.db.title_list, self.db.author_list, self.db.publication_list, self.db.date_list,
             self.db.abstract_list, self.db.doi_list, self.db.web_list]
        # 清空表格
        items = self.excel.get_children()
        [self.excel.delete(item) for item in items]
        # 输出信息
        for i in range(len(self.title_list)):
            self.excel.insert('', i, value=(
                self.title_list[i], self.author_list[i], self.date_list[i], self.publication_list[i],
                self.abstract_list[i], self.doi_list[i],
                self.web_list[i]))

    def download(self):
        d = tk.filedialog.askdirectory()
        dic_name = '\\文献下载'
        if not exists(d + dic_name):
            mkdir(d + dic_name)
        chdir(d + dic_name)

        new = tk.Toplevel()
        win = tk.Text(new)
        win.insert('end', '正在使用sci-hub.tw下载原文.\n下载速度取决于网络质量，耗时较长\n可使用保存功能手动下载所需文献')
        win.pack()

        count = 1
        info = ''
        for doi in self.doi_list:
            url = 'https://www.sci-hub.tw/' + doi

            try:
                req = ur.Request(url, headers=self.db.header)
                html = ur.urlopen(req).read()
                bs = BeautifulSoup(html, 'lxml')
                pdf_link = bs.iframe['src']

                req = ur.Request('https:' + pdf_link, headers=self.db.header)
                response = ur.urlopen(req).read()
                with open(str(count) + '.pdf', 'wb') as f:
                    f.write(response.read())
                info = '第%s篇下载完成' % count
            except Exception:
                info = 'No.%s paper Dowload failed, failed doi is "%s"' % (count, doi)

        win.insert('end', info)
        win.update()
        count += 1

    def file_save(self):
        file_name = filedialog.asksaveasfilename(defaultextension='.xslx', filetypes=[('xlsx', '.xlsx')])
        wb = px.Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = 'PaperInfo'
        for i in range(len(self.columns)):
            sheet1.cell(column=i + 1, row=1, value=self.columns[i])
        for i in range(len(self.title_list)):
            sheet1.cell(column=1, row=i + 2, value=self.title_list[i])
            sheet1.cell(column=2, row=i + 2, value=self.author_list[i])
            sheet1.cell(column=3, row=i + 2, value=self.date_list[i])
            sheet1.cell(column=4, row=i + 2, value=self.publication_list[i])
            sheet1.cell(column=5, row=i + 2, value=self.abstract_list[i])
            sheet1.cell(column=6, row=i + 2, value=self.doi_list[i])
            sheet1.cell(column=7, row=i + 2, value=self.web_list[i])
        wb.save(file_name)
        messagebox.showinfo('保存成功', '文件保存成功' + file_name)

    # 结合excel.bind('<Double-1>', xfunc1)
    # 单击<ButtonRelease-1>
    def open_abs(self, event):
        new = tk.Toplevel()
        new.title('文章信息')
        for item in self.excel.selection():
            item_text = self.excel.item(item, "values")
            text = tk.Text(new, font=self.ft)
            content = 'Title:\n' + item_text[0] + '\n\nPublication\n' + item_text[2] + '\n' + item_text[3] + \
                      '\n\nWeb_Address\n' + item_text[6] + '\n\nAbstract:\n' + item_text[4]
            text.insert(index='insert', chars=content)
            text.grid()

    def search_tip(self, event):
        tk.Label(self.root, text='Searching', width=1, font=self.ft).grid(row=3, column=2, sticky='ew')


if __name__ == '__main__':
    app = Application()
    app.run_it()
